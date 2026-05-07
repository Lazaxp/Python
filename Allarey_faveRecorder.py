import openpyxl as op

wbk = op.Workbook()
sheet = wbk.active

sheet["A1"] = "ID"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Birth Year"
sheet["E1"] = "Age"

print("Welcome to the Favorite People Recorder!")
print("Input Three(3) of your Most Favorite People!")

for x in range(1, 4):
    print(f"Person {x}")
    id = x
    sheet[f"A{x+1}"] = id
    first_name = input("Enter First Name: ")
    sheet[f"B{x+1}"] = first_name
    last_name = input("Enter Last Name: ")
    sheet[f"C{x+1}"] = last_name
    birth_year = input("Enter Birth Year: ")
    sheet[f"D{x+1}"] = birth_year
    age = 2026 - int(birth_year)
    sheet[f"E{x+1}"] = age

wbk.save("Favorite_people.xlsx")

wbk = op.load_workbook("Favorite_people.xlsx")
sheet = wbk.active
for rows in sheet.iter_rows(values_only=True):
    print(rows)
